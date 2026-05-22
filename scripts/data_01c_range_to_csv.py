"""
data_01c_range_to_csv.py

One-shot: fetch a date range from SEIBro and write a clean Excel workbook
matching the on-site layout. Korean labels by default; pass --english to
get a translated version with identical layout.

Usage:
    uv run python scripts/data_01c_range_to_csv.py START END OUT_XLSX [--english]
    # dates as YYYY-MM-DD; range is inclusive on both ends
"""

# %% imports
from __future__ import annotations

import sys
from pathlib import Path
from xml.etree import ElementTree as ET

import httpx
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# %% constants
SEIBRO_URL = "https://seibro.or.kr/websquare/engine/proworks/callServletService.jsp"
SEIBRO_REFERER = (
    "https://seibro.or.kr/websquare/control.jsp"
    "?w2xPath=/IPORTAL/user/ann/BIP_CNST02001V.xml&menuNo=861"
)

# Display order in the on-site .xls export
DISPLAY_MARKETS = ["유로시장", "미국", "일본", "홍콩", "중국", "기타국가", "총합계"]
API_BY_MARKET = {
    "미국": "AA1", "유로시장": "AA2", "일본": "AA3",
    "중국": "AA4", "홍콩": "AA5", "기타국가": "AA_ETC", "총합계": "AA_SUM",
}

# English translations — only labels change, layout & values are identical.
EN_MARKETS = {
    "유로시장": "Eurozone",
    "미국": "US",
    "일본": "Japan",
    "홍콩": "Hong Kong",
    "중국": "China",
    "기타국가": "Other",
    "총합계": "Total",
}
EN_INSTRUMENT = {"주식": "Equity", "채권": "Bond"}
EN_SIDE = {"매도": "Sell", "매수": "Buy"}
EN_REFDATE = "Reference date"
EN_CATEGORY = "Category"
EN_SUMMARY_DATE = "Date"
EN_SUMMARY_EQUITY = "Equity Net Balance"
EN_SUMMARY_DEBT = "Debt Net Balance"


# %% fetch
def fetch_range_xml(start_ymd: str, end_ymd: str) -> str:
    body = (
        '<reqParam action="frsecSetlCusdList" '
        'task="ksd.safe.bip.cnst.Ann.process.AnnSearchPTask">'
        '<PG_START value="1"/><PG_END value="1000"/>'
        f'<ic_start value="{start_ymd}"/><ic_end value="{end_ymd}"/>'
        f'<bDate1 value="{end_ymd}"/>'
        '<S_TYPE value="2"/><GIGAN value="1"/>'
        '<MENU_NO value="861"/>'
        '<CMM_BTN_ABBR_NM value="total_search,openall,print,hwp,word,pdf,seach,xls,"/>'
        '<W2XPATH value="/IPORTAL/user/ann/BIP_CNST02001V.xml"/>'
        '</reqParam>'
    )
    headers = {
        "Content-Type": 'application/xml; charset="UTF-8"',
        "submissionid": "submission_frsecSetlCusdList",
        "Referer": SEIBRO_REFERER,
        "User-Agent": "Mozilla/5.0",
    }
    r = httpx.post(SEIBRO_URL, content=body.encode("utf-8"),
                   headers=headers, timeout=30)
    r.raise_for_status()
    return r.content.decode("utf-8")


def parse_range_rows(xml_text: str) -> list[dict]:
    """Return one row per (date, instrument) with API field names."""
    root = ET.fromstring(xml_text)
    out = []
    for data_el in root.findall("data"):
        result = data_el.find("result")
        if result is None:
            continue
        kv = {child.tag: child.attrib.get("value", "") for child in result}
        out.append(kv)
    return out


def to_csv_lines(rows: list[dict], english: bool = False) -> list[str]:
    """
    Two-row header (markets, then sell/buy), then per-day rows.
    Layout mirrors the on-site Excel export, with one added trailing
    net-balance column:
      net balance = total buy - total sell
    Only the labels change when english=True.
    """
    refdate_h = EN_REFDATE if english else "조회기준일"
    category_h = EN_CATEGORY if english else "구분"
    net_balance_h = "Net balance" if english else "순매수"
    market_label = (lambda m: EN_MARKETS[m]) if english else (lambda m: m)
    sell_label = EN_SIDE["매도"] if english else "매도"
    buy_label = EN_SIDE["매수"] if english else "매수"
    instrument_label = (lambda i: EN_INSTRUMENT.get(i, i)) if english else (lambda i: i)

    h1 = [refdate_h, category_h]
    for m in DISPLAY_MARKETS:
        h1 += [market_label(m), ""]
    h1.append(net_balance_h)
    h2 = ["", ""]
    for _ in DISPLAY_MARKETS:
        h2 += [sell_label, buy_label]
    h2.append("")
    lines = [",".join(h1), ",".join(h2)]

    for kv in rows:
        ymd = kv.get("SETL_DT", "")
        date_str = f"{ymd[:4]}-{ymd[4:6]}-{ymd[6:8]}" if len(ymd) == 8 else ymd
        cells = [date_str, instrument_label(kv.get("SECN_TPCD_NM", ""))]
        for m in DISPLAY_MARKETS:
            code = API_BY_MARKET[m]
            for sfx in ("_1", "_2"):
                cells.append(kv.get(f"{code}{sfx}", ""))
        total_sell = kv.get("AA_SUM_1", "")
        total_buy = kv.get("AA_SUM_2", "")
        cells.append(_fmt_net_balance(total_sell, total_buy))
        lines.append(",".join(cells))
    return lines


def _fmt_net_balance(total_sell: str, total_buy: str) -> str:
    """Compute and format total buy minus total sell as a compact string."""
    try:
        sell = float(total_sell or 0)
        buy = float(total_buy or 0)
    except ValueError:
        return ""
    v = buy - sell
    if v == 0:
        return "0"
    if v == int(v):
        return str(int(v))
    s = f"{v:.10f}".rstrip("0").rstrip(".")
    if s.startswith("0.") and v < 1:
        s = s[1:]
    if s.startswith("-0.") and v > -1:
        s = "-" + s[2:]
    return s


def build_summary_frame(rows: list[dict]) -> pl.DataFrame:
    """Return one row per date with equity and bond net balances."""
    df = pl.DataFrame(rows)
    summary = (
        df.with_columns(
            (pl.col("AA_SUM_2").cast(pl.Float64, strict=False).fill_null(0.0)
             - pl.col("AA_SUM_1").cast(pl.Float64, strict=False).fill_null(0.0)
            ).alias("net_balance"),
            pl.when(pl.col("SECN_TPCD_NM") == "주식")
              .then(pl.lit("Equity"))
              .when(pl.col("SECN_TPCD_NM") == "채권")
              .then(pl.lit("Debt"))
              .otherwise(pl.col("SECN_TPCD_NM"))
              .alias("instrument_en"),
        )
        .select("SETL_DT", "instrument_en", "net_balance")
        .pivot(index="SETL_DT", on="instrument_en", values="net_balance")
        .sort("SETL_DT")
        .rename({"SETL_DT": EN_SUMMARY_DATE, "Equity": EN_SUMMARY_EQUITY, "Debt": EN_SUMMARY_DEBT})
    )
    return summary


def _style_header(ws, row: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        cells = list(col)
        if not cells:
            continue
        letter = cells[0].column_letter
        width = max(len(str(c.value)) if c.value is not None else 0 for c in cells) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), 24)


def build_workbook(rows: list[dict], out_xlsx: Path, english: bool = True) -> Path:
    """Write a two-tab Excel workbook for the email attachment."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws2 = wb.create_sheet("Summary")

    detailed_lines = to_csv_lines(rows, english=english)
    for line in detailed_lines:
        ws1.append(line.split(","))
    _style_header(ws1, 1)
    _style_header(ws1, 2)
    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = ws1.dimensions
    _autosize(ws1)

    summary_df = build_summary_frame(rows)
    ws2.append([EN_SUMMARY_DATE, EN_SUMMARY_EQUITY, EN_SUMMARY_DEBT])
    for row in summary_df.iter_rows(named=True):
        raw_date = str(row.get(EN_SUMMARY_DATE, ""))
        if len(raw_date) == 8 and raw_date.isdigit():
            date_value = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
        else:
            date_value = raw_date
        ws2.append([
            date_value,
            row.get(EN_SUMMARY_EQUITY),
            row.get(EN_SUMMARY_DEBT),
        ])
    _style_header(ws2, 1)
    ws2.freeze_panes = "A2"
    for row in ws2.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
    _autosize(ws2)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx


# %% main
if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--english"]
    english = "--english" in sys.argv
    if len(args) != 3:
        print("usage: data_01c_range_to_csv.py START END OUT_XLSX [--english]",
              file=sys.stderr)
        sys.exit(2)
    start, end, out_xlsx = args[0], args[1], Path(args[2])
    start_ymd = start.replace("-", "")
    end_ymd = end.replace("-", "")
    print(f"[seibro] range {start} -> {end}  (english={english}) ...")
    xml_text = fetch_range_xml(start_ymd, end_ymd)
    rows = parse_range_rows(xml_text)
    print(f"  parsed {len(rows)} rows ({len(rows) // 2} trading days)")
    if not rows:
        print("WARNING: no rows in response", file=sys.stderr)
    out_xlsx = build_workbook(rows, out_xlsx, english=english)
    print(f"  wrote {out_xlsx}  ({out_xlsx.stat().st_size} bytes)")
